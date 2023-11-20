from heapq import merge
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import IconSet, FormatObject, Rule
from openpyxl.chart import PieChart, Reference

filename ="Project2-Solution.xlsx"

wb = load_workbook(filename)
sheet = wb.active
worksheet1 = wb["Carriers and coolers"]
worksheet2 = wb["Int"]

puntaje = 0

#Task1: Group cells A1 to E1 in the first row of the “Int” worksheet. Do not change the alignment of the contents.
mergedCell = worksheet2.merged_cells.ranges
alignm = worksheet2["A1"].alignment.horizontal
lastcell = worksheet2["E1"]
check1 = worksheet2["F1"]
check2 = worksheet2["A2"]

if (type(check1).__name__ != 'MergedCell') and (type(check2).__name__ != 'MergedCell'):
    if (type(mergedCell).__name__ == 'list') and (alignm == None) and (type(lastcell).__name__ == 'MergedCell'):
        puntaje +=1
    else:
        print("No cumplió con TASK 1")
else:
        print("No cumplió con TASK 1")

#Task2: In the worksheet "Carriers and coolers", apply a hyperlink to cells C10, C11, and C12 to cell A4 in the “Int” worksheet.
try:
    hyperl = worksheet1["C10"].hyperlink.location
    hyperl2 = worksheet1["C11"].hyperlink.location
    hyperl3 = worksheet1["C12"].hyperlink.location
    if hyperl and hyperl2 and hyperl3 == "Int!A4":
        puntaje +=1
except AttributeError:
    print("No cumplió con TASK 2")

#Task3: In the worksheet “Carriers and coolers”, apply the 3-flag icon set to the contents of the “Inventory” column.
"""cfCell1 =worksheet1["D4"]
iconset1 = worksheet1.conditional_formatting._cf_rules
first = FormatObject(type='percent', val=0.0)
second =FormatObject(type='percent', val=33.0)
third =FormatObject(type='percent', val=67.0)
iconsetrule = IconSet(iconSet='3Flags', cfvo=[first, second, third])
rule = Rule(type='iconSet',priority =1, iconSet=iconsetrule)
print(rule)
print(type(rule),"\n\n\n")
if iconset1 == rule:
#if(type(iconset1).__name__ == 'iconSet'):
    print("HELOOOOOO!!!\n\n")
#for conditional_formatting in worksheet1.conditional_formatting._cf_rules:
#    for cell_range in conditional_formatting.cells.ranges:
#        if cfCell1.coordinate in cell_range:
#            print("Wiiiuu")
print("AAAAA",iconset1)
print("TYPEEEEE",type(iconset1))
inicio = worksheet2["A6"].value
final = worksheet2["D13"].value
if inicio == 10 and final == 1750:
    puntaje +=1
else:
    print("No cumplió con TASK 3")

#Task4: In the worksheet, expand the chart data range to include the rest of the rows in the table
dataRange = Reference(worksheet1,min_col=3,min_row=4,max_col=19)
piechart = PieChart()
chartnew = worksheet1._charts
print(chartnew)
print(type(chartnew))
#if position == "Boat":
#    puntaje +=1
#else:
#    print("No cumplió con TASK 4")
"""
#Task 5. Hide the “Inv” worksheet.
try:
    wbt5 = load_workbook("InventoryReport.xlsx")
    if (type(wbt5).__name__) == "Workbook":
        puntaje +=1
except FileNotFoundError:
        print("No cumplió con TASK 5")

print("\nPuntaje de Proyecto A:", (puntaje),"/5")

wb.save(filename)